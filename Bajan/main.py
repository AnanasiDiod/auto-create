import openpyxl as xl
import os
from math import ceil


def calculate_pdv_1_pr(width, height, cost_met, cost_cut, cost_bend, cost_vata, cost_axis, num_axis,
                       cost_strip, cost_screw, extra_cost, markup_box, cost_work, cost_drive, markup_drive, additional,
                       ms, cost_warm_cable, cost_warm_work, quantity):

    markup_box = markup_box / 100 + 1
    markup_drive = markup_drive / 100 + 1
    extra_cost = extra_cost / 100
    # Расчёт длины реза. Каждая строка - отдельная деталь
    len_cut = height * (4 + 4 + 4 + 4) + \
        width * (4 + 4 + 4 + 4) + \
        (1076.346 + 37.699 +  # Крышка привода
         # Боковина лопатки вертикаль
         (134.959 - 48 * 2)*2 + 35.5 + 49.2 +
         # Боковина лопатки горизонт
         (134.172 - 82.4 * 2) * 2 +
         # Половина лопатки
         (-16 * 2 - 58.4 * 2 + 12.566 + 15.39*2)*2 +
         # Площадка
         869.712 + 21.99 * 4 + 56.549 +
         # Профиль 1
         (918.296 - 17.6 * 2 + 15.39 * 16)*2 + 20.42 * 4 + 56.549 +
         # Профиль 2
         (1278.676 + 46 * 2 + 15.39 * 15 + 48.274 * 4) * 2 + 70.324 +
         # Боковина
         (205.92 - 4 * 2 + 15.39 * 4) * 2 + 34.558 + 56.549 +
         # Поддержка оси
         211.36 + 15.39 * 4 + 12.3 * 4 +
         # Уголок
         (77.27 - 98 + 15.39 * 3) * 2 +
         # Шайбы 2мм
         95.819 * 4)
    # Учёт отверстий в половине лопатки и в боковинах
    nh = (((height - 66) // 120) // 2 + 1) * 2
    nw = ((width - 98) // 120) + 2
    len_cut += (nh + nw) * 8 * 15.39
    # Расчёт длины гиба
    len_bend = ((height - 4) * 4 * 2 +  # Боковина
                # Боковина лопатки
                (height - 18) * 2 * 2 +
                # Боковина лопатки горизонтальная
                (width - 54.4) * 2 * 2 +
                # Крыщка привода
                494.8 +
                # Площадка
                864.8 +
                # Поддержка оси
                222.72 +
                # Профиль 1
                (312.4 * 2 + (height - 7.6) * 4) * 2 +
                # Профиль 2
                (width - 1) * 2 * 2 +
                # Уголок
                (width - 60) * 2)
    # Расчёт площади металла
    area_met_korp = (103 * 2 + 391 * 2) * height + (389 * 2 + 27.64 * 2) * width - \
        (4 * 103 * 2 - 55 * 391 * 2 - 56 * 389 * 2 +
         38 * 27.64) + 198 * 314 + 153 * 291
    area_met_lopatka = (height - 12) * (width - 54) * 2 + \
        (55.69 * 2 + 0) * height + (55.69 * 2 + 0) * width - \
        (18 * 55.69 * 2 + 54 * 55.69 * 2) + 50 * 55.68
    # длина ленты и кабеля
    len_strip = ((height - 18) + (width - 54.4)) * 2 / 1000
    len_warm_cable = (height + width - 8) * 2 / 1000 + (0.16 + 0.08) * 2 + 1
    val_warm_cable = len_warm_cable * cost_warm_cable
    # площадь лопатки
    area_lopatka = (width - 54.4) * (height - 18)
    # стоимость оси
    len_axis = 0.14
    # учёт работ и материалов для морозостойкости
    val_met_warm = ((height - 4 + width - 4) * 2 * 102.96 + 70.04 * 68.37 * 4
                    + 176.94 * 263.94 + 111 * (259.24 + 491.64)) * cost_met / 1000000
    val_cut_warm = (767.81 + height * 4 + (351.86 + width * 2)
                    * 2 + 365.8 * 4 + 976 + 834.73 + 1544.57) * cost_cut / 1000
    val_bend_warm = ((height - 4 + width - 4) * 4 * 2 + (37 +
                     68.37 * 2) * 4 + (225.38 + 137.19) * 2 + 111 * 5) * cost_bend / 1000
    val_vata_warm = (176.94 * 263.94 + 111 * (259.24 + 491.64) +
                     (height - 4 + width - 4) * 39.14 * 2) * cost_vata / 10000000
    len_axsis_warm = 0.04
    val_warm_work = 0
    # Учёт наличия ребра
    if not (width < 500 and height < 500 or height < 200):
        len_cut += 548.58 + width * 2 - 84.4 * 2 + 15.39 * 9
        area_met_lopatka += (width - 84) * 198
        len_bend += (width - 84.4) * 4
    # Расчёт "голой" стоимости металла, стоимости реза и стоимости гиба
    if ms:
        len_axis += len_axsis_warm
        val_warm_work = cost_warm_work
    area_lopatka /= 1000000
    val_bend = cost_bend * len_bend / 1000
    val_cut = len_cut * cost_cut / 1000
    area_met_lopatka /= 1000000
    area_met_korp /= 1000000
    val_met = (area_met_lopatka + area_met_korp) * cost_met
    val_vata = area_lopatka * cost_vata
    val_axis = len_axis * cost_axis * num_axis

    res = {'width': width,
           'height': height,
           'quad': area_met_lopatka + area_met_korp,
           'cost_met': cost_met,
           'val_met': val_met,
           'val_cut': val_cut,
           'val_bend': val_bend,
           'vata': val_vata,
           'axis': val_axis,
           'strip': cost_strip * len_strip,
           'screw': cost_screw,
           'extra': extra_cost * (val_met + val_vata + val_cut + val_bend + val_axis + cost_strip * len_strip + cost_screw),
           'markup_metall': markup_box,
           'work': cost_work,
           'ms': ms,
           'warm_met': val_met_warm,
           'warm_cut': val_cut_warm,
           'warm_bend': val_bend_warm,
           'warm_vata': val_vata_warm,
           'warm_cable': val_warm_cable,
           'warm_work': val_warm_work,
           'drive': cost_drive,
           'markup_drive': markup_drive,
           'add': additional,
           'quantity': quantity}
    res['total'] = ((res['val_met'] + res['val_cut'] + res['val_bend'] + res['axis'] + res['vata'] + res['strip'] + res['screw'] + res['extra']) * markup_box +
                    res['work'] + res['drive'] * markup_drive + res['add'])
    if ms:
        res['total'] += (res['warm_met'] + res['warm_cut'] + res['warm_bend'] +
                         res['warm_vata'] + res['warm_cable']) * markup_box + res['warm_work']
    res['total'] *= quantity

    # print('pdv_1_pr')

    return res

def calculate_pdv_1_pr_ei150(width, height, cost_met, cost_cut, cost_bend, cost_vata, cost_axis, num_axis,
                       cost_strip, cost_screw, extra_cost, markup_box, cost_work, cost_drive, markup_drive, additional,
                       ms, cost_warm_cable, cost_warm_work, quantity):

    markup_box = markup_box / 100 + 1
    markup_drive = markup_drive / 100 + 1
    extra_cost = extra_cost / 100
    # Расчёт длины реза. Каждая строка - отдельная деталь
    len_cut = height * (4 + 4 + 4 + 4) + \
        width * (4 + 4 + 4 + 4) + \
        (1076.346 + 37.699 +  # Крышка привода
         # Боковина лопатки вертикаль
         (134.959 - 48 * 2)*2 + 35.5 + 49.2 +
         # Боковина лопатки горизонт
         (134.172 - 82.4 * 2) * 2 +
         # Половина лопатки
         (-16 * 2 - 58.4 * 2 + 12.566 + 15.39*2)*2 +
         # Площадка
         869.712 + 21.99 * 4 + 56.549 +
         # Профиль 1
         (918.296 - 17.6 * 2 + 15.39 * 16)*2 + 20.42 * 4 + 56.549 +
         # Профиль 2
         (1278.676 + 46 * 2 + 15.39 * 15 + 48.274 * 4) * 2 + 70.324 +
         # Боковина
         (205.92 - 4 * 2 + 15.39 * 4) * 2 + 34.558 + 56.549 +
         # Поддержка оси
         211.36 + 15.39 * 4 + 12.3 * 4 +
         # Уголок
         (77.27 - 98 + 15.39 * 3) * 2 +
         # Шайбы 2мм
         95.819 * 4 +
         # Дополнительный профиль горизонтальный
         (1349.34 + width * 2) * 2 +
         # Дополнительный профиль вертикальный
         (668.66 + height * 2) * 2)
    # Учёт отверстий в половине лопатки и в боковинах
    nh = (((height - 66) // 120) // 2 + 1) * 2
    nw = ((width - 98) // 120) + 2
    len_cut += (nh + nw) * 8 * 15.39
    # Расчёт длины гиба
    len_bend = ((height - 4) * 4 * 2 +  # Боковина
                # Боковина лопатки
                (height - 18) * 2 * 2 +
                # Боковина лопатки горизонтальная
                (width - 54.4) * 2 * 2 +
                # Крыщка привода
                494.8 +
                # Площадка
                864.8 +
                # Поддержка оси
                222.72 +
                # Профиль 1
                (312.4 * 2 + (height - 7.6) * 4) * 2 +
                # Профиль 2
                (width - 1) * 2 * 2 +
                # Уголок
                (width - 60) * 2 +
                # Дополнительный профиль горизонтальный
                (width - 1) * 4 +
                # Дополнительный профиль вертикальный
                (height - 7.6) * 4 + 92.4 * 4)
    # Расчёт площади металла
    area_met_korp = (103 * 2 + 391 * 2) * height + (389 * 2 + 27.64 * 2) * width - \
        (4 * 103 * 2 - 55 * 391 * 2 - 56 * 389 * 2 +
         38 * 27.64) + 198 * 314 + 153 * 291 + 168.96 * (width + 56) * 2 + 170.56 * (height + 55.28) * 2
    area_met_lopatka = (height - 12) * (width - 54) * 2 + \
        (55.69 * 2 + 0) * height + (55.69 * 2 + 0) * width - \
        (18 * 55.69 * 2 + 54 * 55.69 * 2) + 50 * 55.68
    # длина ленты и кабеля
    len_strip = ((height - 18) + (width - 54.4)) * 2 / 1000
    len_warm_cable = (height + width - 8) * 2 / 1000 + (0.16 + 0.08) * 2 + 1
    val_warm_cable = len_warm_cable * cost_warm_cable
    # площадь лопатки
    area_lopatka = (width - 54.4) * (height - 18)
    # стоимость оси
    len_axis = 0.14
    # учёт работ и материалов для морозостойкости
    val_met_warm = ((height - 4 + width - 4) * 2 * 102.96 + 70.04 * 68.37 * 4
                    + 176.94 * 263.94 + 111 * (259.24 + 491.64)) * cost_met / 1000000
    val_cut_warm = (767.81 + height * 4 + (351.86 + width * 2)
                    * 2 + 365.8 * 4 + 976 + 834.73 + 1544.57) * cost_cut / 1000
    val_bend_warm = ((height - 4 + width - 4) * 4 * 2 + (37 +
                     68.37 * 2) * 4 + (225.38 + 137.19) * 2 + 111 * 5) * cost_bend / 1000
    val_vata_warm = (176.94 * 263.94 + 111 * (259.24 + 491.64) +
                     (height - 4 + width - 4) * 39.14 * 2) * cost_vata / 10000000
    len_axsis_warm = 0.04
    val_warm_work = 0
    # Учёт наличия ребра
    if not (width < 500 and height < 500 or height < 200):
        len_cut += 548.58 + width * 2 - 84.4 * 2 + 15.39 * 9
        area_met_lopatka += (width - 84) * 198
        len_bend += (width - 84.4) * 4
    # Расчёт "голой" стоимости металла, стоимости реза и стоимости гиба
    if ms:
        len_axis += len_axsis_warm
        val_warm_work = cost_warm_work
    area_lopatka /= 1000000
    val_bend = cost_bend * len_bend / 1000
    val_cut = len_cut * cost_cut / 1000
    area_met_lopatka /= 1000000
    area_met_korp /= 1000000
    val_met = (area_met_lopatka + area_met_korp) * cost_met
    val_vata = (area_lopatka + (height + width) * 2 * 20 / 1000000) * cost_vata
    val_axis = len_axis * cost_axis * num_axis

    res = {'width': width,
           'height': height,
           'quad': area_met_lopatka + area_met_korp,
           'cost_met': cost_met,
           'val_met': val_met,
           'val_cut': val_cut,
           'val_bend': val_bend,
           'vata': val_vata,
           'axis': val_axis,
           'strip': cost_strip * len_strip,
           'screw': cost_screw,
           'extra': extra_cost * (val_met + val_vata + val_cut + val_bend + val_axis + cost_strip * len_strip + cost_screw),
           'markup_metall': markup_box,
           'work': cost_work,
           'ms': ms,
           'warm_met': val_met_warm,
           'warm_cut': val_cut_warm,
           'warm_bend': val_bend_warm,
           'warm_vata': val_vata_warm,
           'warm_cable': val_warm_cable,
           'warm_work': val_warm_work,
           'drive': cost_drive,
           'markup_drive': markup_drive,
           'add': additional,
           'quantity': quantity}
    res['total'] = ((res['val_met'] + res['val_cut'] + res['val_bend'] + res['axis'] + res['vata'] + res['strip'] + res['screw'] + res['extra']) * markup_box +
                    res['work'] + res['drive'] * markup_drive + res['add'])
    if ms:
        res['total'] += (res['warm_met'] + res['warm_cut'] + res['warm_bend'] +
                         res['warm_vata'] + res['warm_cable']) * markup_box + res['warm_work']
    res['total'] *= quantity

    # print('pdv_1_pr_ei150')

    return res


def calculate_pdv_1_kr(width, height, cost_met, cost_cut, cost_bend, cost_vata, cost_axis, num_axis,
                       cost_strip, cost_screw, extra_cost, markup_box, cost_work, cost_drive, markup_drive, additional,
                       ms, cost_warm_cable, cost_warm_work, quantity):

    markup_box = markup_box / 100 + 1
    markup_drive = markup_drive / 100 + 1
    extra_cost = extra_cost / 100
    # Расчёт длины реза. Каждая строка - отдельная деталь
    if int(height) > int(width):
        width = height
    else:
        height = width
    len_cut = ((3.141592 * (width - 1) + 10.3) * 2 + 400 * 2 + 192.58 +  # обечайка
               1076.346 + 37.699 +  # крышка привода
               # лопатка
               (3.141592 * (width - 8) + (12 + 1.3) * 2 * int(width * 3 / 100 + 3) + 15.4 * 9 + 20.4 * 2) * 2 +
               1263.8 +  # площадка
               306.42 +  # поддержка оси
               367.56 +  # поддержка оси квадрат
               385.09 +  # поддержка оси центр
               (width - 50 + 56.7) * 2 + 15.4 * 6 + 49.2 +  # поддержка оси центр 2
               # ребро
               ((20 + 3.141592 * (width - 10) / 2) * 2 + 22.07 * int(width * 3 / 100 + 3)) * 2 +
               150.2)  # уголок
    len_bend = (3.141592 * (width - 10) * 2 +  # ребро
                15 +  # уголок
                96.2 * 4 + 109.98 +  # крышка привода
                (172 + 110 + 30 + 16.2 * 2) * 2 +  # площадка
                20 * 4 + 200 * 2 + 15 + (width - 50) * 2)  # всё остальное
    # Расчёт площади металла
    area_met_korp = (3.141592 * (width - 1) + 10.3) * 400 + \
        314.48 * 197.64 + 152.88 * 250.56 + 15 * 47.64
    area_met_lopatka = 2 * 3.141592 * ((width/2 - 4) ** 2) + 3.141592 * (
        width - 10) * 36.37 * 2 + 56.74 * 40 * 2 + 56.74 * 200 + 56.74 * (width - 50)
    # длина ленты и кабеля
    len_strip = 3.141592 * (width - 8) / 1000
    len_warm_cable = len_strip + (0.16 + 0.08) * 2 + 1
    val_warm_cable = len_warm_cable * cost_warm_cable
    # площадь лопатки
    area_lopatka = 3.141592 * ((width/2 - 4) ** 2)
    # стоимость оси
    len_axis = 0.14
    # учёт работ и материалов для морозостойкости
    val_met_warm = (176.94 * 263.94 + 111 *
                    (259.24 + 491.64)) * cost_met / 1000000
    val_cut_warm = (976 + 834.73 + 1544.57) * cost_cut / 1000
    val_bend_warm = ((225.38 + 137.19) * 2 + 111 * 5) * cost_bend / 1000
    val_vata_warm = (176.94 * 263.94 + 111 * (259.24 + 491.64)
                     ) * cost_vata / 10000000
    len_axsis_warm = 0.04
    val_warm_work = 0
    # Расчёт "голой" стоимости металла, стоимости реза и стоимости гиба
    if ms:
        len_axis += len_axsis_warm
        val_warm_work = cost_warm_work
    area_lopatka /= 1000000
    val_bend = cost_bend * len_bend / 1000
    val_cut = len_cut * cost_cut / 1000
    area_met_lopatka /= 1000000
    area_met_korp /= 1000000
    val_met = (area_met_lopatka + area_met_korp) * cost_met
    val_vata = area_lopatka * cost_vata
    val_axis = len_axis * cost_axis * num_axis

    res = {'width': width,
           'height': height,
           'quad': area_met_lopatka + area_met_korp,
           'cost_met': cost_met,
           'val_met': val_met,
           'val_cut': val_cut,
           'val_bend': val_bend,
           'vata': val_vata,
           'axis': val_axis,
           'strip': cost_strip * len_strip,
           'screw': cost_screw,
           'extra': extra_cost * (val_met + val_vata + val_cut + val_bend + val_axis + cost_strip * len_strip + cost_screw),
           'markup_metall': markup_box,
           'work': cost_work,
           'ms': ms,
           'warm_met': val_met_warm,
           'warm_cut': val_cut_warm,
           'warm_bend': val_bend_warm,
           'warm_vata': val_vata_warm,
           'warm_cable': val_warm_cable,
           'warm_work': val_warm_work,
           'drive': cost_drive,
           'markup_drive': markup_drive,
           'add': additional,
           'quantity': quantity}
    res['total'] = ((res['val_met'] + res['val_cut'] + res['val_bend'] + res['axis'] + res['vata'] + res['strip'] + res['screw'] + res['extra']) * markup_box +
                    res['work'] + res['drive'] * markup_drive + res['add'])
    if ms:
        res['total'] += (res['warm_met'] + res['warm_cut'] + res['warm_bend'] +
                         res['warm_vata'] + res['warm_cable']) * markup_box + res['warm_work']
    res['total'] *= quantity

    # print('pdv_1_kr')

    return res


def calculate_pdv_2_s_ei(width, height, cost_met, cost_cut, cost_bend, cost_vata, cost_axis, num_axis,
                         cost_strip, cost_screw, extra_cost, markup_box, cost_work, cost_drive, markup_drive, additional,
                         ms, cost_warm_cable, cost_warm_work, quantity):

    markup_box = markup_box / 100 + 1
    markup_drive = markup_drive / 100 + 1
    extra_cost = extra_cost / 100
    # Расчёт длины реза
    nh = ((height - 18) // 120) + 1
    nw = ((width - 93) // 120) + 1
    len_cut = ((156.13 * 2 + 100.53 + 20.4 + 62.83 + 103.67) * 2 +  # Рычаг 1
               # Рычаг 2
               (109.5 * 2 + 2 * 2 + 62.83 + 40.84) * 2 +
               # Боковина для оси
               89.49 * 2 + 20 * 2 + 15.7 + 30.79 + 53.4 +
               # Боковина лопатки вертикальная
               ((22.1 * 4 + 23.25 + (height - 73)) * 2 + 35.5 + nh * 15.39) * 2 +
               # Боковиана лопатки горизонтальная
               ((20 * 2 + 27 + (width - 121)) * 2 + nw * 15.39) * 2 +
               # Половина лопатки зад
               ((height - 22) + (width - 97)) * \
               2 + 12.57 + (nh + nw) * 2 * 15.39 + 6 * 15.39 +
               # Половина лопатки перед
               ((height - 36) + (width - 97)) * 2 + \
               12.57 + (nh + nw) * 2 * 15.39 + 3 * 15.39 + (21 + 5.5) * 2 * 2 +
               # Профиль 1
               (163.79 + 18.44 + 5.42 + 43.8 + 24.13 + 10 + (height + 43.06) + 10 + 24.13 + 43.8 + \
                5.24 + 16.44 + 163.79 + 16.15 + 14.15 * 2 + (height - 29.81) + 14.15 * 2 + 12.57 + 15.7 + \
                8 * 15.4 + 28.27) * 2 + 30.78 + 30.63 +
               # Профиль 2
               188.08 + 55.43 + 2.22 + \
               (width - 143.14) + 2.22 + 55.43 + 188.08 + 23 + 24.13 + 10 + (width + 26.86) + 10 + 24.13 + 62.83 + 6 * 15.39 + 2 * 28.27 +
               # Профиль 3
               182.9 + (width - 15) + 182.9 + 23 + 24.13 + 10 + (width + 26.86) + 10 + 24.13 + 47.12 + 6 * 15.4 + 138.23 +
               # Ребро
               (198.29 + width - 123) * 2 + 9 * 15.39 + (29.1 * 2 + 20 + 31.4 + 15.39) * 2 +
               # Боковина
               (97.52 + (height - 23.8) + 64.48 + 1 + 33.04 +
                (height - 28.8) + 6.28 + 4 * 15.39 + 34.56) * 2 +
               # Поддержка оси
               89.5 * 2 + 20 * 2 + 15.7 + 30.79 + 53.4
               )
    # Расчёт длины гиба
    len_bend = ((width - 93) * 2 * 2 +  # боковина лопатки горизонтальная
                # боковина лопатки вертикальная
                (height - 43) * 2 * 2 +
                # Боковина
                (height - 23.8) * 3 * 2 +
                # Профиль 1
                (height - 25.8 + height - 30.4) * 2 +
                # Профиль 2 и 3
                (width - 5) * 2 * 2 +
                # Поддержка оси
                30 * 4 +
                # Ребро
                (width - 123) * 4 +
                # Рычаг 1
                32
                )
    # Расчёт площади металла
    area_met_korp = + 99.521 * (height - 23.8) * 2 + \
        (height + 167.2) * 230.38 * 2 + (width + 51) * (234.28 + 229.1) + \
        30 * 99.5 + 32 * 188 + 22 * 130
    area_met_lopatka = (width - 93) * (height - 18) * 2 + (width - 93) * \
        55.68 * 2 + (height - 43) * 55.68 * 2 + (width - 123) * 198.288
    # длина ленты и кабеля
    len_strip = ((height - 43) + (width - 93)) * 2 / 1000
    len_warm_cable = (height + width - 8) * 2 / 1000 + (0.16 + 0.08) * 2 + 1
    val_warm_cable = len_warm_cable * cost_warm_cable
    # площадь лопатки
    area_lopatka = (height - 43) * (width - 93)
    # стоимость оси
    len_axis = 0.14
    # учёт работ и материалов для морозостойкости
    val_met_warm = ((height - 4 + width - 4) * 2 * 102.96 + 70.04 * 68.37 * 4
                    + 176.94 * 263.94 + 111 * (259.24 + 491.64)) * cost_met / 1000000
    val_cut_warm = (767.81 + height * 4 + (351.86 + width * 2)
                    * 2 + 365.8 * 4 + 976 + 834.73 + 1544.57) * cost_cut / 1000
    val_bend_warm = ((height - 4 + width - 4) * 4 * 2 + (37 +
                     68.37 * 2) * 4 + (225.38 + 137.19) * 2 + 111 * 5) * cost_bend / 1000
    val_vata_warm = (176.94 * 263.94 + 111 * (259.24 + 491.64) +
                     (height - 4 + width - 4) * 39.14 * 2) * cost_vata / 10000000
    len_axsis_warm = 0.04
    val_warm_work = 0
    # Расчёт "голой" стоимости металла, стоимости реза и стоимости гиба
    if ms:
        len_axis += len_axsis_warm
        val_warm_work = cost_warm_work
    area_lopatka /= 1000000
    val_bend = cost_bend * len_bend / 1000
    val_cut = len_cut * cost_cut / 1000
    area_met_lopatka /= 1000000
    area_met_korp /= 1000000
    val_met = (area_met_lopatka + area_met_korp) * cost_met
    val_vata = area_lopatka * cost_vata
    val_axis = len_axis * cost_axis * num_axis

    res = {'width': width,
           'height': height,
           'quad': area_met_lopatka + area_met_korp,
           'cost_met': cost_met,
           'val_met': val_met,
           'val_cut': val_cut,
           'val_bend': val_bend,
           'vata': val_vata,
           'axis': val_axis,
           'strip': cost_strip * len_strip,
           'screw': cost_screw,
           'extra': extra_cost * (val_met + val_vata + val_cut + val_bend + val_axis + cost_strip * len_strip + cost_screw),
           'markup_metall': markup_box,
           'work': cost_work,
           'ms': ms,
           'warm_met': val_met_warm,
           'warm_cut': val_cut_warm,
           'warm_bend': val_bend_warm,
           'warm_vata': val_vata_warm,
           'warm_cable': val_warm_cable,
           'warm_work': val_warm_work,
           'drive': cost_drive,
           'markup_drive': markup_drive,
           'add': additional,
           'quantity': quantity}
    res['total'] = ((res['val_met'] + res['val_cut'] + res['val_bend'] + res['axis'] + res['vata'] + res['strip'] + res['screw'] + res['extra']) * markup_box +
                    res['work'] + res['drive'] * markup_drive + res['add'])
    if ms:
        res['total'] += (res['warm_met'] + res['warm_cut'] + res['warm_bend'] +
                         res['warm_vata'] + res['warm_cable']) * markup_box + res['warm_work']
    res['total'] *= quantity

    # print('pdv_2_s_ei')

    return res


def calculate_pdv_2_s(width, height, cost_met, cost_cut, cost_bend, cost_vata, cost_axis, num_axis,
                      cost_strip, cost_screw, extra_cost, markup_box, cost_work, cost_drive, markup_drive, additional,
                      ms, cost_warm_cable, cost_warm_work, quantity):

    markup_box = markup_box / 100 + 1
    markup_drive = markup_drive / 100 + 1
    extra_cost = extra_cost / 100
    # Расчёт длины реза
    len_cut = ((156.13 * 2 + 100.53 + 20.4 + 62.83 + 103.67) * 2 +  # Рычаг 1
               # Рычаг 2
               (120 * 2 + 2 * 2 + 62.83 + 40.84) * 2 +
               # Боковина для оси
               89.5 * 2 + 20 * 2 + 15.7 + 30.79 + 53.4 +
               # Лопатка
               (width - 49.74) * 2 + 18.51 * 2 * 4 + (height - 93.67) * 2 + (14.14 + 20 + 14.14) * 2 + 17.93 * 2 + 30.79 + 34.56 +
               # Профиль 1
               (168.89 + 18.44 + 5.42 + 46.8 + 27.13 + 10 + (height + 42.86) + 10 + 27.13 + 46.8 + \
                5.24 + 18.44 + 168.89 + 16.15 + 14.15 * 2 + (height - 30.02) + 14.15 * 2 + 12.57 + 15.7 + \
                6 * 15.4 + 12.57 + 28.27) * 2 + 30.78 + 30.63 +
               # Профиль 2
               197.78 * 2 + (width - 9) + 23 + 24.13 + 10 + \
               width + 26.86 + 10 + 24.13 + 6.28 + 31.42 + 56.55 + 6 * 15.4 +
               # Профиль 3
               187.38 + (width - 9) + 187.38 + 23 + 24.13 + 10 + (width + 26.86) + 10 + 24.13 + 31.42 + 6.3 + 56.55 + 6 * 15.4 + 138.23 +
               # Ребро
               (100.29 + width - 98) * 2 + 30.79 + 40.84 + (25 + 22 + 25 + 6 + 25.13) * 2 +
               # Полоска горизонтальная
               (width - 93 + 32.5 + 6.28) * 2 +
               # Полоска вертикальная
               (height - 78 + 32.5 + 6.28) * 2 * 2
               )
    # Учёт отверстий в лопатке и полосках
    nh = int((270 + (height - 400))//120)
    nw = int((288.2 + (width - 400))//120)
    len_cut += (nh * 2 + nw) * 2 * 15.39
    # Расчёт длины гиба
    len_bend = ((width - 49.74) * 2 + (height - 35.74) * 2 +  # Лопатка
                # Профиль 1
                (168.89 * 2 + (height - 30.02) * 2) * 2 +
                # Профиль 2 и 3
                + (width - 5) * 2 * 2 +
                # Боковина для оси
                30 * 4 +
                # Ребро
                (width - 98) * 4 + 22 * 2 +
                # Рычаг 1
                32
                )
    # Расчёт площади металла
    area_met_korp = 231.48 * (height + 67) * 2 + \
        (width + 51) * 241 + (width + 51) * 230.6 + 99.5 * \
        30 + (height - 78) * 36.5 * 2 + (width - 93) * \
        36.5 + 100.29 * (width - 98) + 32 * 188.2 * 2 + 140 * 22 * 2
    area_met_lopatka = (height + 1.28) * (width + 7.28)
    # длина ленты и кабеля
    len_strip = ((width - 49.74) + (height - 35.74)) * 2 / 1000
    len_warm_cable = len_strip + (0.16 + 0.08) * 2 + 1
    val_warm_cable = len_warm_cable * cost_warm_cable
    # площадь лопатки
    area_lopatka = (width - 49.74) * (height - 35.74) / 1000000
    # стоимость оси
    len_axis = 0.14
    # учёт работ и материалов для морозостойкости
    val_met_warm = ((height - 4 + width - 4) * 2 * 102.96 + 70.04 * 68.37 * 4
                    + 176.94 * 263.94 + 111 * (259.24 + 491.64)) * cost_met / 1000000
    val_cut_warm = (767.81 + height * 4 + (351.86 + width * 2)
                    * 2 + 365.8 * 4 + 976 + 834.73 + 1544.57) * cost_cut / 1000
    val_bend_warm = ((height - 4 + width - 4) * 4 * 2 + (37 +
                     68.37 * 2) * 4 + (225.38 + 137.19) * 2 + 111 * 5) * cost_bend / 1000
    val_vata_warm = (176.94 * 263.94 + 111 * (259.24 + 491.64) +
                     (height - 4 + width - 4) * 39.14 * 2) * cost_vata / 10000000
    len_axsis_warm = 0.04
    val_warm_work = 0
    # Расчёт "голой" стоимости металла, стоимости реза и стоимости гиба
    if ms:
        len_axis += len_axsis_warm
        val_warm_work = cost_warm_work
    area_lopatka /= 1000000
    val_bend = cost_bend * len_bend / 1000
    val_cut = len_cut * cost_cut / 1000
    area_met_lopatka /= 1000000
    area_met_korp /= 1000000
    val_met = (area_met_lopatka + area_met_korp) * cost_met
    val_vata = area_lopatka * cost_vata
    val_axis = len_axis * cost_axis * num_axis

    res = {'width': width,
           'height': height,
           'quad': area_met_lopatka + area_met_korp,
           'cost_met': cost_met,
           'val_met': val_met,
           'val_cut': val_cut,
           'val_bend': val_bend,
           'vata': val_vata,
           'axis': val_axis,
           'strip': cost_strip * len_strip,
           'screw': cost_screw,
           'extra': extra_cost * (val_met + val_vata + val_cut + val_bend + val_axis + cost_strip * len_strip + cost_screw),
           'markup_metall': markup_box,
           'work': cost_work,
           'ms': ms,
           'warm_met': val_met_warm,
           'warm_cut': val_cut_warm,
           'warm_bend': val_bend_warm,
           'warm_vata': val_vata_warm,
           'warm_cable': val_warm_cable,
           'warm_work': val_warm_work,
           'drive': cost_drive,
           'markup_drive': markup_drive,
           'add': additional,
           'quantity': quantity}
    res['total'] = ((res['val_met'] + res['val_cut'] + res['val_bend'] + res['axis'] + res['vata'] + res['strip'] + res['screw'] + res['extra']) * markup_box +
                    res['work'] + res['drive'] * markup_drive + res['add'])
    if ms:
        res['total'] += (res['warm_met'] + res['warm_cut'] + res['warm_bend'] +
                         res['warm_vata'] + res['warm_cable']) * markup_box + res['warm_work']
    res['total'] *= quantity

    # print('pdv_2_s')

    return res


def calculate_pdv_2_k(width, height, cost_met, cost_cut, cost_bend, cost_vata, cost_axis, num_axis,
                      cost_strip, cost_screw, extra_cost, markup_box, cost_work, cost_drive, markup_drive, additional,
                      ms, cost_warm_cable, cost_warm_work, quantity):

    markup_box = markup_box / 100 + 1
    markup_drive = markup_drive / 100 + 1
    extra_cost = extra_cost / 100
    # Расчёт длины реза
    len_cut = (210.39 + width * 2 + height * 2 +  # Лопатка
               1014.23 +  # Площадка
               741 + width * 2 +  # Полоса низ
               1039.57 + height * 2 +  # профиль 1 под привод
               932.13 + height * 2 +  # профиль 1 обычный
               1642.9 + width * 2 +  # профиль 2 верх
               1655.36 + width * 2 +  # профиль 2 низ
               (38.16 + height * 2) * 2 +  # уголок бок
               68.16 + width * 2 +  # уголок верх низ
               332.39 +  # боковина для оси
               334.87 + width * 2 +  # ребро_3
               (558.51 + 339.51) * 2)  # рычаг 1 и 2
    len_bend = ((width - 11.74 + height - 15.34) * 2 +  # Лопатка
                (212.4 + 220) * 2 +  # Площадка
                (182.4 + (height - 7.6) * 2) * 2 * 2 +  # профиль 1
                (width * 2 + 55) * 2 * 2 +  # профиль 2
                (height - 55) * 2 * 2 +  # уголок бок
                (width - 40) * 2 +  # уголок верх низ
                120 +  # боковина для оси
                (width - 60) * 4 + 44 +  # ребро_3
                64 * 2)  # рычаг 1
    # Расчёт площади металла
    area_met_korp = (231.48 * (height + 67) * 2 + 290.56 * 152.88 +
                     42 * (width + 2) + 260.56 * (height + 55.28) * 2 +
                     258.96 * (width + 56) * 2 + 55.28 * (height - 55) * 2 +
                     (width - 40) * 55.28 + 99.49 * 30 + (width - 60) * 100.29 +
                     178.07 * 32 * 2 + 22 * 140 * 2)
    area_met_lopatka = (height + 1.28) * (width + 7.28)
    # длина ленты и кабеля
    len_strip = (width - 11.74 + height - 15.34) * 2 / 1000
    len_warm_cable = len_strip + (0.16 + 0.08) * 2 + 1
    val_warm_cable = len_warm_cable * cost_warm_cable
    # площадь лопатки
    area_lopatka = (width - 11.74) * (height - 15.34) / 1000000
    # стоимость оси
    len_axis = 0.14
    # учёт работ и материалов для морозостойкости
    val_met_warm = ((height - 4 + width - 4) * 2 * 102.96 + 70.04 * 68.37 * 4
                    + 176.94 * 263.94 + 111 * (259.24 + 491.64)) * cost_met / 1000000
    val_cut_warm = (767.81 + height * 4 + (351.86 + width * 2)
                    * 2 + 365.8 * 4 + 976 + 834.73 + 1544.57) * cost_cut / 1000
    val_bend_warm = ((height - 4 + width - 4) * 4 * 2 + (37 +
                     68.37 * 2) * 4 + (225.38 + 137.19) * 2 + 111 * 5) * cost_bend / 1000
    val_vata_warm = (176.94 * 263.94 + 111 * (259.24 + 491.64) +
                     (height - 4 + width - 4) * 39.14 * 2) * cost_vata / 10000000
    len_axsis_warm = 0.04
    val_warm_work = 0
    # Расчёт "голой" стоимости металла, стоимости реза и стоимости гиба
    if ms:
        len_axis += len_axsis_warm
        val_warm_work = cost_warm_work
    area_lopatka /= 1000000
    val_bend = cost_bend * len_bend / 1000
    val_cut = len_cut * cost_cut / 1000
    area_met_lopatka /= 1000000
    area_met_korp /= 1000000
    val_met = (area_met_lopatka + area_met_korp) * cost_met
    val_vata = area_lopatka * cost_vata
    val_axis = len_axis * cost_axis * num_axis

    res = {'width': width,
           'height': height,
           'quad': area_met_lopatka + area_met_korp,
           'cost_met': cost_met,
           'val_met': val_met,
           'val_cut': val_cut,
           'val_bend': val_bend,
           'vata': val_vata,
           'axis': val_axis,
           'strip': cost_strip * len_strip,
           'screw': cost_screw,
           'extra': extra_cost * (val_met + val_vata + val_cut + val_bend + val_axis + cost_strip * len_strip + cost_screw),
           'markup_metall': markup_box,
           'work': cost_work,
           'ms': ms,
           'warm_met': val_met_warm,
           'warm_cut': val_cut_warm,
           'warm_bend': val_bend_warm,
           'warm_vata': val_vata_warm,
           'warm_cable': val_warm_cable,
           'warm_work': val_warm_work,
           'drive': cost_drive,
           'markup_drive': markup_drive,
           'add': additional,
           'quantity': quantity}
    res['total'] = ((res['val_met'] + res['val_cut'] + res['val_bend'] + res['axis'] + res['vata'] + res['strip'] + res['screw'] + res['extra']) * markup_box +
                    res['work'] + res['drive'] * markup_drive + res['add'])
    if ms:
        res['total'] += (res['warm_met'] + res['warm_cut'] + res['warm_bend'] +
                         res['warm_vata'] + res['warm_cable']) * markup_box + res['warm_work']
    res['total'] *= quantity

    # print('pdv_2_k')

    return res


def calculate_pdv_2_k_ei(width, height, cost_met, cost_cut, cost_bend, cost_vata, cost_axis, num_axis,
                         cost_strip, cost_screw, extra_cost, markup_box, cost_work, cost_drive, markup_drive, additional,
                         ms, cost_warm_cable, cost_warm_work, quantity):

    markup_box = markup_box / 100 + 1
    markup_drive = markup_drive / 100 + 1
    extra_cost = extra_cost / 100
    # Расчёт длины реза
    len_cut = ((210.39 + width * 2 + height * 2) * 2 +  # Лопатка
               1014.23 +  # Площадка
               741 + width * 2 +  # Полоса низ
               1039.57 + height * 2 +  # профиль 1 под привод
               932.13 + height * 2 +  # профиль 1 обычный
               1642.9 + width * 2 +  # профиль 2 верх
               1655.36 + width * 2 +  # профиль 2 низ
               (38.16 + height * 2) * 2 +  # уголок бок
               68.16 + width * 2 +  # уголок верх низ
               332.39 +  # боковина для оси
               334.87 + width * 2 +  # ребро_3
               (558.51 + 339.51) * 2)  # рычаг 1 и 2
    len_bend = ((width - 11.74 + height - 15.34) * 2 * 2 +  # Лопатка
                (212.4 + 220) * 2 +  # Площадка
                (182.4 + (height - 7.6) * 2) * 2 * 2 +  # профиль 1
                (width * 2 + 55) * 2 * 2 +  # профиль 2
                (height - 55) * 2 * 2 +  # уголок бок
                (width - 40) * 2 +  # уголок верх низ
                120 +  # боковина для оси
                (width - 60) * 4 + 44 +  # ребро_3
                64 * 2)  # рычаг 1
    # Расчёт площади металла
    area_met_korp = (231.48 * (height + 67) * 2 + 290.56 * 152.88 +
                     42 * (width + 2) + 260.56 * (height + 55.28) * 2 +
                     258.96 * (width + 56) * 2 + 55.28 * (height - 55) * 2 +
                     (width - 40) * 55.28 + 99.49 * 30 + (width - 60) * 100.29 +
                     178.07 * 32 * 2 + 22 * 140 * 2)
    area_met_lopatka = (height + 1.28) * (width + 7.28)
    # длина ленты и кабеля
    len_strip = (width - 11.74 + height - 15.34) * 2 / 1000
    len_warm_cable = len_strip + (0.16 + 0.08) * 2 + 1
    val_warm_cable = len_warm_cable * cost_warm_cable
    # площадь лопатки
    area_lopatka = (width - 11.74) * (height - 15.34) * 2 / 1000000
    # стоимость оси
    len_axis = 0.14
    # учёт работ и материалов для морозостойкости
    val_met_warm = ((height - 4 + width - 4) * 2 * 102.96 + 70.04 * 68.37 * 4
                    + 176.94 * 263.94 + 111 * (259.24 + 491.64)) * cost_met / 1000000
    val_cut_warm = (767.81 + height * 4 + (351.86 + width * 2)
                    * 2 + 365.8 * 4 + 976 + 834.73 + 1544.57) * cost_cut / 1000
    val_bend_warm = ((height - 4 + width - 4) * 4 * 2 + (37 +
                     68.37 * 2) * 4 + (225.38 + 137.19) * 2 + 111 * 5) * cost_bend / 1000
    val_vata_warm = (176.94 * 263.94 + 111 * (259.24 + 491.64) +
                     (height - 4 + width - 4) * 39.14 * 2) * cost_vata / 10000000
    len_axsis_warm = 0.04
    val_warm_work = 0
    # Расчёт "голой" стоимости металла, стоимости реза и стоимости гиба
    if ms:
        len_axis += len_axsis_warm
        val_warm_work = cost_warm_work
    area_lopatka /= 1000000
    val_bend = cost_bend * len_bend / 1000
    val_cut = len_cut * cost_cut / 1000
    area_met_lopatka /= 1000000
    area_met_korp /= 1000000
    val_met = (area_met_lopatka + area_met_korp) * cost_met
    val_vata = area_lopatka * cost_vata
    val_axis = len_axis * cost_axis * num_axis

    res = {'width': width,
           'height': height,
           'quad': area_met_lopatka + area_met_korp,
           'cost_met': cost_met,
           'val_met': val_met,
           'val_cut': val_cut,
           'val_bend': val_bend,
           'vata': val_vata,
           'axis': val_axis,
           'strip': cost_strip * len_strip,
           'screw': cost_screw,
           'extra': extra_cost * (val_met + val_vata + val_cut + val_bend + val_axis + cost_strip * len_strip + cost_screw),
           'markup_metall': markup_box,
           'work': cost_work,
           'ms': ms,
           'warm_met': val_met_warm,
           'warm_cut': val_cut_warm,
           'warm_bend': val_bend_warm,
           'warm_vata': val_vata_warm,
           'warm_cable': val_warm_cable,
           'warm_work': val_warm_work,
           'drive': cost_drive,
           'markup_drive': markup_drive,
           'add': additional,
           'quantity': quantity}
    res['total'] = ((res['val_met'] + res['val_cut'] + res['val_bend'] + res['axis'] + res['vata'] + res['strip'] + res['screw'] + res['extra']) * markup_box +
                    res['work'] + res['drive'] * markup_drive + res['add'])
    if ms:
        res['total'] += (res['warm_met'] + res['warm_cut'] + res['warm_bend'] +
                         res['warm_vata'] + res['warm_cable']) * markup_box + res['warm_work']
    res['total'] *= quantity

    # print('pdv_2_k_ei')

    return res


def calculate_pdv_2_lk(width, height, cost_met, cost_cut, cost_bend, cost_vata, cost_axis, num_axis,
                       cost_strip, cost_screw, extra_cost, markup_box, cost_work, cost_drive, markup_drive, additional,
                       ms, cost_warm_cable, cost_warm_work, quantity):

    markup_box = markup_box / 100 + 1
    markup_drive = markup_drive / 100 + 1
    extra_cost = extra_cost / 100
    np = int(height / 220) + 1  # количество лопаток
    hp = (height - 7.5 * (np + 1)) / np  # высота лопатки
    # количество горизонтальных отверстий в лопатках
    nw = ceil((width - 114.4)/125) + 1
    # Расчёт длины реза. Каждая строка - отдельная деталь
    len_cut = (((height - 5 + 102.96) * 2 + 15.4 * 4) * 2 + 34.56 * (np * 2 - 1) + 56.55 +  # Боковина + под привод
               # боковина лопатки + под привод
               ((67.4 + 15.4 * 2 + hp - 36) * 2 + 35.5) * 2 * np + 13.7 +
               # боковина лопатки горизонтальная
               ((67.4 + nw * 15.4 + width - 82.4) * 2) * 2 * np +
               1114.03 +  # крышка привода
               1014.23 +  # площадка
               # половина лопатки + под привод
               ((width - 60.4 + hp - 1) * 2 + 18.85 + (4 + nw * 2) * 15.4) * np + 2 * 15.4 +
               # половина лопатки с отгибами + под привод
               ((116.05 + hp - 60.215 + width - 60.4 + 19.48) * 2 + (4 + nw * 2) * 15.4) * np + 2 * 15.4 +
               # профиль 1 + под привод
               (960.184 + 2 * height) * 2 + 56.55 + 4 * 20.42 +
               (1701.29 + width * 2) * 2 - 1.32 +  # профиль 2 верх и низ
               (25.46 + width * 2) * 2 +  # уголок верх низ
               (-300.17 + height * 2 + 20.42 * (np - 2)) * 2 +  # ребро
               315.7)  # поддержка оси

    len_bend = ((width - 60) * 2 +  # уголок
                (height - 5) * 4 * 2 +  # боковина
                hp * 2 * 2 * np +  # боковина лопатки
                # боковина лопатки горизонтальная
                (width - 54.4) * 2 * 2 * np +
                494.8 +  # крышка привода
                (212.4 + 110 * 2) * 2 +  # площадка
                57.215 * 2 * np +  # половина лопатки с отгибами
                (212.4 * 2 + (height - 7.6) * 2 * 2) * 2 +  # профиль 1
                (width - 1) * 2 * 2 * 2 +  # профиль 2
                100)  # поддержка оси
    # Расчёт площади металла
    area_met_korp = 102.98 * (height - 5) * 2 + 314.48 * 197.64 + 290.56 * 152.88 + (height + 55.28) * \
        290.56 * 2 + 288.96 * (width + 56) * 2 + 27.64 * \
        (width - 38) + 22 * (np * hp + 10) * 2
    area_met_lopatka = 55.69 * np * \
        (hp + width - 54.4) * 2 + (width - 54.4 +
                                   width + 32.48) * hp * np + 55.68 * 50
    # длина ленты и кабеля
    len_strip = (height - 5 + np * (width - 54.4)) * 2 / 1000
    len_warm_cable = (height + width - 8) * 2 / 1000 + (0.16 + 0.08) * 2 + 1
    val_warm_cable = len_warm_cable * cost_warm_cable
    # площадь лопатки
    area_lopatka = np * hp * (width - 54.4) / 1000000
    # стоимость оси
    len_axis = 0.14
    # учёт работ и материалов для морозостойкости
    val_met_warm = ((height - 4 + width - 4) * 2 * 102.96 + 70.04 * 68.37 * 4
                    + 176.94 * 263.94 + 111 * (259.24 + 491.64)) * cost_met / 1000000
    val_cut_warm = (767.81 + height * 4 + (351.86 + width * 2)
                    * 2 + 365.8 * 4 + 976 + 834.73 + 1544.57) * cost_cut / 1000
    val_bend_warm = ((height - 4 + width - 4) * 4 * 2 + (37 +
                     68.37 * 2) * 4 + (225.38 + 137.19) * 2 + 111 * 5) * cost_bend / 1000
    val_vata_warm = (176.94 * 263.94 + 111 * (259.24 + 491.64) +
                     (height - 4 + width - 4) * 39.14 * 2) * cost_vata / 10000000
    len_axsis_warm = 0.04
    val_warm_work = 0
    # Расчёт "голой" стоимости металла, стоимости реза и стоимости гиба
    if ms:
        len_axis += len_axsis_warm
        val_warm_work = cost_warm_work
    area_lopatka /= 1000000
    val_bend = cost_bend * len_bend / 1000
    val_cut = len_cut * cost_cut / 1000
    area_met_lopatka /= 1000000
    area_met_korp /= 1000000
    val_met = (area_met_lopatka + area_met_korp) * cost_met
    val_vata = area_lopatka * cost_vata
    val_axis = len_axis * cost_axis * num_axis

    res = {'width': width,
           'height': height,
           'quad': area_met_lopatka + area_met_korp,
           'cost_met': cost_met,
           'val_met': val_met,
           'val_cut': val_cut,
           'val_bend': val_bend,
           'vata': val_vata,
           'axis': val_axis,
           'strip': cost_strip * len_strip,
           'screw': cost_screw,
           'extra': extra_cost * (val_met + val_vata + val_cut + val_bend + val_axis + cost_strip * len_strip + cost_screw),
           'markup_metall': markup_box,
           'work': cost_work,
           'ms': ms,
           'warm_met': val_met_warm,
           'warm_cut': val_cut_warm,
           'warm_bend': val_bend_warm,
           'warm_vata': val_vata_warm,
           'warm_cable': val_warm_cable,
           'warm_work': val_warm_work,
           'drive': cost_drive,
           'markup_drive': markup_drive,
           'add': additional,
           'quantity': quantity}
    res['total'] = ((res['val_met'] + res['val_cut'] + res['val_bend'] + res['axis'] + res['vata'] + res['strip'] + res['screw'] + res['extra']) * markup_box +
                    res['work'] + res['drive'] * markup_drive + res['add'])
    if ms:
        res['total'] += (res['warm_met'] + res['warm_cut'] + res['warm_bend'] +
                         res['warm_vata'] + res['warm_cable']) * markup_box + res['warm_work']
    res['total'] *= quantity

    # print('pdv_2_lk')

    return res


def calculate_pdv_2_ls(width, height, cost_met, cost_cut, cost_bend, cost_vata, cost_axis, num_axis,
                       cost_strip, cost_screw, extra_cost, markup_box, cost_work, cost_drive, markup_drive, additional,
                       ms, cost_warm_cable, cost_warm_work, quantity):

    markup_box = markup_box / 100 + 1
    markup_drive = markup_drive / 100 + 1
    extra_cost = extra_cost / 100
    np = int(height / 220.8) + 1  # количество лопаток
    hp = (height - 7.5 * (np + 1)) / np  # высота лопатки
    # количество горизонтальных отверстий в лопатках
    nw = ceil((width - 273.4)/125) + 1
    nh = ceil((height - 42.443)/125) + 1
    # Расчёт длины реза. Каждая строка - отдельная деталь
    len_cut = ((319.41 + height * 2 - 34.56 + (np - 2) * 34.56) + 22 + 18.85 * 2 +  # Боковина + под привод
               # Боковина лопатки горизонтальная
               (-287.04 + width * 2 + (nw - 2) * 2 * 15.4) * 2 +
               # Боковина лопатки + под привод
               (147.92 + hp * 2) * 2 * np + 13.7 +
               322.14 +  # Поддержка оси
               # Половина лопатки + под привод
               ((hp - 1 + width - 263.4) * 2 + 8 * 15.4 + (nw - 2) * 2 * 15.4) * np + 2 * 15.4 +
               # Половина лопатки с отгибами + под привод
               (323.22 + width * 2 + hp * 2 + (nw - 2) * 2 * 15.4) * np + 2 * 15.4 +
               # Профиль 1 + под привод
               (894.15 + height * 2) * 2 + (nh + 4) * 15.4 +
               (755.89 + width * 2) * 2 + 138.227 +  # Профиль 2 верх и низ
               969.82 + height * 2 + nh * 15.4 +  # Стенка
               (-292.54 + width * 2) * 2 +  # Уголок верх низ
               (-300.17 + height * 2 + 20.42 * (np - 2)) * 2)  # ребро
    len_bend = ((width - 197) * 2 +  # уголок
                (height - 8.6) * 4 * 2 +  # боковина
                hp * 2 * 2 * np +  # боковина лопатки
                # боковина лопатки горизонтальная
                (width - 213.4) * 2 * 2 * np +
                57.215 * 2 * np +  # половина лопатки с отгибами
                (213.6 * 2 + (height - 9.6) * 2) * 2 +  # профиль 1
                (width - 5) * 2 * 2 +  # профиль 2
                100 +  # поддержка оси
                (height - 9.6) * 3 + 174.16 * 2)  # Стенка
    # Расчёт площади металла
    area_met_korp = 102.96 * (height - 8.6) * 2 + 276.18 * (height + 83.4) * 2 + 282.28 * (
        width + 51) * 2 + 333.72 * (height + 28.68) + 27.64 * (width - 197) * 2 + 22 * (np * hp + 10) * 2
    area_met_lopatka = 55.69 * np * \
        (hp + width - 213.4) * 2 + (width - 213.4 +
                                    width + 126.6) * hp * np + 55.68 * 50
    # длина ленты и кабеля
    len_strip = (height - 8.6 + np * (width - 213.4)) * 2 / 1000
    len_warm_cable = (height + width - 8) * 2 / 1000 + (0.16 + 0.08) * 2 + 1
    val_warm_cable = len_warm_cable * cost_warm_cable
    # площадь лопатки
    area_lopatka = np * hp * (width - 213.4) / 1000000
    # стоимость оси
    len_axis = 0.14
    # учёт работ и материалов для морозостойкости
    val_met_warm = ((height - 4 + width - 4) * 2 * 102.96 + 70.04 * 68.37 * 4
                    + 176.94 * 263.94 + 111 * (259.24 + 491.64)) * cost_met / 1000000
    val_cut_warm = (767.81 + height * 4 + (351.86 + width * 2)
                    * 2 + 365.8 * 4 + 976 + 834.73 + 1544.57) * cost_cut / 1000
    val_bend_warm = ((height - 4 + width - 4) * 4 * 2 + (37 +
                     68.37 * 2) * 4 + (225.38 + 137.19) * 2 + 111 * 5) * cost_bend / 1000
    val_vata_warm = (176.94 * 263.94 + 111 * (259.24 + 491.64) +
                     (height - 4 + width - 4) * 39.14 * 2) * cost_vata / 10000000
    len_axsis_warm = 0.04
    val_warm_work = 0
    # Расчёт "голой" стоимости металла, стоимости реза и стоимости гиба
    if ms:
        len_axis += len_axsis_warm
        val_warm_work = cost_warm_work
    area_lopatka /= 1000000
    val_bend = cost_bend * len_bend / 1000
    val_cut = len_cut * cost_cut / 1000
    area_met_lopatka /= 1000000
    area_met_korp /= 1000000
    val_met = (area_met_lopatka + area_met_korp) * cost_met
    val_vata = area_lopatka * cost_vata
    val_axis = len_axis * cost_axis * num_axis

    res = {'width': width,
           'height': height,
           'quad': area_met_lopatka + area_met_korp,
           'cost_met': cost_met,
           'val_met': val_met,
           'val_cut': val_cut,
           'val_bend': val_bend,
           'vata': val_vata,
           'axis': val_axis,
           'strip': cost_strip * len_strip,
           'screw': cost_screw,
           'extra': extra_cost * (val_met + val_vata + val_cut + val_bend + val_axis + cost_strip * len_strip + cost_screw),
           'markup_metall': markup_box,
           'work': cost_work,
           'ms': ms,
           'warm_met': val_met_warm,
           'warm_cut': val_cut_warm,
           'warm_bend': val_bend_warm,
           'warm_vata': val_vata_warm,
           'warm_cable': val_warm_cable,
           'warm_work': val_warm_work,
           'drive': cost_drive,
           'markup_drive': markup_drive,
           'add': additional,
           'quantity': quantity}
    res['total'] = ((res['val_met'] + res['val_cut'] + res['val_bend'] + res['axis'] + res['vata'] + res['strip'] + res['screw'] + res['extra']) * markup_box +
                    res['work'] + res['drive'] * markup_drive + res['add'])
    if ms:
        res['total'] += (res['warm_met'] + res['warm_cut'] + res['warm_bend'] +
                         res['warm_vata'] + res['warm_cable']) * markup_box + res['warm_work']
    res['total'] *= quantity

    # print('pdv_2_ls')

    return res


def extract_to_excel(res=list(), name=str(), path=str()):
    wb = xl.Workbook()
    ws = wb.active
    for i in range(len(res)):
        ws['A' + str(1 + i * 3)] = res[i]['name']
        ws['A' + str(2 + i * 3)] = 'Сторона А'
        ws['A' + str(3 + i * 3)] = str(res[i]['width']).replace('.', ',')
        ws['B' + str(2 + i * 3)] = 'Сторона B'
        ws['B' + str(3 + i * 3)] = str(res[i]['height']).replace('.', ',')
        ws['C' + str(2 + i * 3)] = 'Квадратура клапана'
        ws['C' + str(3 + i * 3)] = str(res[i]['quad']).replace('.', ',')
        ws['D' + str(2 + i * 3)] = 'Стоимость металла [м2]'
        ws['D' + str(3 + i * 3)] = str(res[i]['cost_met']).replace('.', ',')
        ws['E' + str(2 + i * 3)] = 'Стоимость железа за коробку клапана'
        ws['E' + str(3 + i * 3)] = str(res[i]['val_met']).replace('.', ',')
        ws['F' + str(2 + i * 3)] = 'Стоимость резки'
        ws['F' + str(3 + i * 3)] = str(res[i]['val_cut']).replace('.', ',')
        ws['G' + str(2 + i * 3)] = 'Стоимость гиба'
        ws['G' + str(3 + i * 3)] = str(res[i]['val_bend']).replace('.', ',')
        ws['H' + str(2 + i * 3)] = 'Стоимость ваты'
        ws['H' + str(3 + i * 3)] = str(res[i]['vata']).replace('.', ',')
        ws['I' + str(2 + i * 3)] = 'Стоимость оси(-ей)'
        ws['I' + str(3 + i * 3)] = str(res[i]['axis']).replace('.', ',')
        ws['J' + str(2 + i * 3)] = 'Стоимость ленты'
        ws['J' + str(3 + i * 3)] = str(res[i]['strip']).replace('.', ',')
        ws['K' + str(2 + i * 3)] = 'Стоимость крепежа'
        ws['K' + str(3 + i * 3)] = str(res[i]['screw']).replace('.', ',')
        ws['L' + str(2 + i * 3)] = 'Дополнительные расходы на металл'
        ws['L' + str(3 + i * 3)] = str(res[i]['extra']).replace('.', ',')
        ws['M' + str(2 + i * 3)] = 'Наценка на корпус клапана'
        ws['M' + str(3 + i * 3)
           ] = str(round((res[i]['markup_metall'] - 1) * 100, 2)) + '%'
        ws['N' + str(2 + i * 3)] = 'Стоимость работ'
        ws['N' + str(3 + i * 3)] = str(res[i]['work']).replace('.', ',')
        ws['O' + str(2 + i * 3)] = 'Стоимость привода'
        ws['O' + str(3 + i * 3)] = str(res[i]['drive']).replace('.', ',')
        ws['P' + str(2 + i * 3)] = 'Наценка на привод'
        ws['P' + str(3 + i * 3)
           ] = str(round((res[i]['markup_drive'] - 1) * 100, 2)).replace('.', ',')
        ws['Q' + str(2 + i * 3)] = 'Стоимость утепления'
        ws['Q' + str(3 + i * 3)] = str(res[i]['warm_met'] + res[i]['warm_cut'] +
                                       res[i]['warm_bend'] + res[i]['warm_vata'] + res[i]['warm_cable'] + res[i]['warm_work']).replace('.', ',')
        ws['R' + str(2 + i * 3)] = 'Дополнительно'
        ws['R' + str(3 + i * 3)] = str(res[i]['add']).replace('.', ',')
        ws['S' + str(2 + i * 3)] = 'Количество'
        ws['S' + str(3 + i * 3)] = str(res[i]['quantity']).replace('.', ',')
        ws['T' + str(2 + i * 3)] = 'Итоговая стоимость'
        ws['T' + str(3 + i * 3)] = str(res[i]['total']).replace('.', ',')
        # print(path + '\\' + res[i]['name'] + '.xlsx')
    
    fname = name.split('.')[0] + ".xlsx"
    wb.save(os.path.join(path, fname))
    return fname


def extract_to_kp(res=list(), name=str(), path=str()):
    wb = xl.Workbook()
    ws = wb.active
    for i in range(len(res)):
        ws['A' + str(1 + i)] = str(res[i]['name']) + \
            " - " + str(int(res[i]['quantity'])) + 'шт'
        ws['B' + str(1 + i)] = str(int(res[i]['total']))
    wb.save(os.path.join(path, name + ".xlsx"))
    return name + '.xlsx'
